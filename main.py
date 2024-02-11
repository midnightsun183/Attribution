import os.path
import tkinter as tk
from tkinter import *
# from tkinter import * #filedialog, messagebox, Menu, Toplevel, Tk, Frame, Label, Button, Checkbutton, BooleanVar
import subprocess  # Import subprocess
from tkinter import messagebox, filedialog

import openpyxl
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import copy
import pickle
import threading
from tkinter import Tk, PhotoImage


# from PIL import Image, ImageTk

class App:
    def __init__(self, root):

        self.cancel_button = None
        self.process_csv_button = None
        self.select_file_button = None
        self.settings_frame = None
        self.main_frame = None
        self.root = root

        # Threads
        self.process_thread = None  # To store the thread running the subprocess
        self.current_subprocess = None  # To store the current running subprocess

        # User Settings
        self.settings_directory = "Assets"
        self.settings_filename = "app_settings.pkl"
        self.settings_path = os.path.join(self.settings_directory, self.settings_filename)
        # Load settings or use default values
        settings = self.load_settings()
        self.dark_mode_enabled = tk.BooleanVar(value=settings.get('dark_mode_enabled', True))
        self.portfolio_only = tk.BooleanVar(value=settings.get('portfolio_only', False))

        self.apply_window_size(settings.get('window_size', (800, 400)))  # Apply window size settings

        self.text_color = '#000000'
        self.bg_color = '#FFFFFF'
        self.normal_bg = "#263D42"
        self.normal_fg = "#FFFFFF"
        self.select_bg = "#347B98"
        self.select_fg = "#FFD700"

        self.fill = None
        self.font = None
        self.border = None
        self.protection = None
        self.alignment = None
        self.number_format = None
        self.value = None
        self.has_style = None
        self.file_path = ""  # Initialize file_path to store the selected file path
        self.initialize_gui()
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)  # Bind the custom method to the window's close event

    def initialize_gui(self):
        self.root.title("Ariel Application")
        # Set the path to your icon

        # icon_path = os.path.join("Assets", "logo.ico")
        # self.root.iconbitmap(icon_path)
        # Continue setting up your UI components

        self.create_menu()
        self.setup_frames()  # Set up frames

    def create_menu(self):
        # Menu Bar
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        # File Menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Combine Excel Files", command=self.combine_excel_files)
        file_menu.add_separator()
        file_menu.add_command(label="Quit", command=self.close_app)
        menu_bar.add_cascade(label="File", menu=file_menu)

        # Settings menu
        settings_menu = tk.Menu(menu_bar, tearoff=0)
        settings_menu.add_command(label="Options", command=self.show_settings)
        menu_bar.add_cascade(label="Settings", menu=settings_menu)

        # Navigation menu
        navigation_menu = tk.Menu(menu_bar, tearoff=0)
        navigation_menu.add_command(label="Main Menu", command=self.display_main_widgets)
        navigation_menu.add_command(label="Settings", command=self.show_settings)
        menu_bar.add_cascade(label="Navigate", menu=navigation_menu)

        # Help Menu
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="Help", command=lambda: self.show_message("Help", "This is the help message."))
        help_menu.add_command(label="About", command=lambda: self.show_message("About", "This is the About message."))
        menu_bar.add_cascade(label="Help", menu=help_menu)

    def close_app(self):
        # Call cancel_process to ensure any running subprocess is terminated
        self.cancel_process()
        # Check if there's a need to do any other cleanup before exiting
        # ~~~ Other operations ~~~

        # Finally, destroy the root window to close the app
        self.root.destroy()

    def setup_frames(self):
        self.main_frame = tk.Frame(self.root)
        self.main_frame.place(relwidth=1.0, relheight=1.0)
        self.settings_frame = tk.Frame(self.root)

        self.display_main_widgets()
        self.apply_theme()

    def apply_theme(self):
        self.bg_color = "#263D42" if self.dark_mode_enabled.get() else "#FFFFFF"
        self.text_color = "#FFFFFF" if self.dark_mode_enabled.get() else "#000000"
        self.update_widgets_theme()

    def update_widgets_theme(self):
        for frame in (self.main_frame, self.settings_frame):
            frame.config(bg=self.bg_color)
            for widget in frame.winfo_children():
                if isinstance(widget, (tk.Label, tk.Button, tk.Checkbutton)):
                    widget.config(bg=self.bg_color,
                                  fg=self.text_color,
                                  activebackground=self.bg_color,
                                  activeforeground=self.text_color,
                                  )

    def display_main_widgets(self):
        # Initially, all buttons are visible; visibility adjustments will be made when starting/stopping subprocesses
        self.clear_frame(self.settings_frame)
        self.settings_frame.place_forget()  # Hide settings_frame
        # self.clear_frame(self.main_frame)  # Assuming this clears the main_frame
        self.main_frame.place(relwidth=1.0, relheight=1.0)  # Adjust if necessary

        # Store button references
        self.select_file_button = tk.Button(self.main_frame, text="Select File", command=self.select_file,
                                            bg=self.bg_color, activebackground=self.bg_color,
                                            activeforeground=self.text_color, fg=self.text_color)

        self.process_csv_button = tk.Button(self.main_frame, text="Process CSV",
                                            command=self.call_external_script_in_thread,
                                            bg=self.bg_color, activebackground=self.bg_color,
                                            activeforeground=self.text_color, fg=self.text_color)

        self.cancel_button = tk.Button(self.main_frame, text="Cancel", command=self.cancel_process,
                                       bg=self.bg_color, activebackground=self.bg_color,
                                       activeforeground=self.text_color, fg=self.text_color)

        # Configure and place the output_display Text widget
        self.output_display = tk.Text(self.main_frame, wrap=tk.WORD, height=10, width=50,
                                      bg=self.bg_color, fg=self.text_color)
        # Adjust the relwidth and relheight to make the display area larger
        self.output_display.place(relx=0.05, rely=0.7, relwidth=0.9, relheight=0.2)

        # Place buttons

        self.select_file_button.place(relx=0.4, rely=0.4, relwidth=0.2, relheight=0.1)
        self.process_csv_button.place(relx=0.4, rely=0.5, relwidth=0.2, relheight=0.1)
        self.cancel_button.place(relx=0.4, rely=0.6, relwidth=0.2, relheight=0.1)

        self.output_display.place(relx=0.15, rely=0.2, relwidth=0.75, relheight=0.1)

    def display_output(self, text):
        """Displays the given text in the output_display widget."""
        self.output_display.insert(tk.END, text)
        self.output_display.see(tk.END)  # Scroll to the end of the text widget to show the latest output

    def show_buttons(self):
        self.select_file_button.place(relx=0.4, rely=0.4, relwidth=0.2, relheight=0.1)
        self.process_csv_button.place(relx=0.4, rely=0.5, relwidth=0.2, relheight=0.1)

    def show_settings(self):
        # Your settings display code, modified to include Save Settings button
        # self.clear_frame(self.main_frame)
        self.clear_frame(self.settings_frame)
        self.settings_frame.place(relwidth=1.0, relheight=1.0)

        tk.Label(self.settings_frame, text="Settings", font=('Arial', 24), bg=self.bg_color, fg=self.text_color).place(relx=0.4, rely=0.4, relwidth=0.2, relheight=0.1)

        tk.Checkbutton(self.settings_frame, text="Dark Mode", variable=self.dark_mode_enabled, command=self.apply_theme,
                       activebackground=self.bg_color,
                       activeforeground=self.text_color,
                       bg=self.bg_color,
                       fg=self.text_color,
                       selectcolor='white' if not self.dark_mode_enabled.get() else self.bg_color).place(relx=0.4, rely=0.4, relwidth=0.4, relheight=0.1)

        tk.Button(self.settings_frame, text="Back to main", command=self.display_main_widgets,
                  bg=self.bg_color,
                  activebackground=self.bg_color, activeforeground=self.text_color,
                  fg=self.text_color).place(relx=0.4, rely=0.6, relwidth=0.2, relheight=0.1)

        tk.Button(self.settings_frame, text="Save Settings", command=self.save_settings,
                  bg=self.bg_color,
                  activebackground=self.bg_color, activeforeground=self.text_color,
                  fg=self.text_color).place(relx=0.4, rely=0.8, relwidth=0.2, relheight=0.1)

        tk.Checkbutton(self.settings_frame, text="Portfolio Only", variable=self.portfolio_only,
                       activebackground=self.bg_color,
                       activeforeground=self.text_color,
                       bg=self.bg_color,
                       fg=self.text_color,
                       selectcolor='white' if not self.dark_mode_enabled.get() else self.bg_color).place(relx=0.4, rely=0.9, relwidth=0.2, relheight=0.1)

    def save_settings(self):
        # Ensure a minimum size of 200x200
        width = max(self.root.winfo_width(), 200)
        height = max(self.root.winfo_height(), 200)
        settings = {
            'dark_mode_enabled': self.dark_mode_enabled.get(),
            'portfolio_only': self.portfolio_only.get(),
            'window_size': (width, height)
        }
        if not os.path.exists(self.settings_directory):
            os.makedirs(self.settings_directory)
        with open(self.settings_path, 'wb') as f:
            pickle.dump(settings, f)
        messagebox.showinfo("Settings Saved", "Your settings have been saved.")

    def load_settings(self):
        if os.path.exists(self.settings_path):
            with open(self.settings_path, 'rb') as f:
                return pickle.load(f)
        return {}  # Return an empty dict if the settings file doesn't exist

    def apply_window_size(self, size):
        self.root.geometry(f"{size[0]}x{size[1]}")

    def clear_frame(self, frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def show_message(self, title, message):
        messagebox.showinfo(title, message)

    def select_file(self):
        file_path = tk.filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("Excel File", "*xlsx")])
        if file_path:
            self.file_path = file_path  # Store the selected file path
        else:
            messagebox.showinfo("Information", "No file was selected.")

    # Ensure call_external_script_in_thread starts the process thread properly
    def call_external_script_in_thread(self):
        if self.file_path:  # Ensure a file has been selected
            _, file_extension = os.path.splitext(self.file_path)
            script_name = "ABC.py" if file_extension.lower() == '.xlsx' else "ABC.py"
            try:
                # Hide buttons when the subprocess starts
                self.root.after(0, self.select_file_button.place_forget)
                self.root.after(0, self.process_csv_button.place_forget)

                # Handle starting the subprocess within a thread
                if self.current_subprocess is None or self.current_subprocess.poll() is not None:  # Not running
                    self.process_thread = threading.Thread(target=self.run_subprocess, args=(script_name, self.file_path))
                    self.process_thread.start()
                else:
                    messagebox.showinfo("Process Running", "A process is already running. Please wait or cancel it before starting a new one.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to start the process: {e}")
        else:
            messagebox.showinfo("Information", "Please select a file first.")

    def run_subprocess(self, script_name, file_path):
        try:
            self.current_subprocess = subprocess.Popen(
                ["python", script_name, self.file_path, str(self.portfolio_only.get())],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)

            # Read output line by line
            while True:
                output_line = self.current_subprocess.stdout.readline()
                if output_line == '' and self.current_subprocess.poll() is not None:
                    break
                if output_line:
                    self.root.after(0, self.display_output, output_line)
            self.current_subprocess.wait()

        except subprocess.CalledProcessError as e:
            self.root.after(0, messagebox.showerror, "Error", f"Failed to process the file: {e}")

            self.root.after(0, self.show_buttons)
        except Exception as e:
            self.root.after(0, messagebox.showerror, "Unexpected Error", f"An unexpected error occurred: {e}")
            self.root.after(0, self.show_buttons)
        finally:
            # Reset the current_subprocess to None after it completes or fails
            self.current_subprocess = None
            self.root.after(0, self.show_buttons)

    def cancel_process(self):
        # Reset the file path
        self.file_path = ""
        # Attempt to terminate the subprocess if it's running
        if self.current_subprocess and self.current_subprocess.poll() is None:
            self.current_subprocess.terminate()
            self.current_subprocess.wait()  # Wait for the subprocess to terminate
            self.current_subprocess = None
            messagebox.showinfo("Cancelled", "The process has been cancelled.")
        else:
            messagebox.showinfo("Cancelled", "No process was running.")


    def copy_style(self, source_cell, target_cell):
        # Copy cell style, including font, fill, border, alignment, and number format
        if source_cell.has_style:
            # target_cell.font = source_cell.font.copy()
            # target_cell.border = source_cell.border.copy()
            # target_cell.fill = source_cell.fill.copy()
            # target_cell.number_format = source_cell.number_format
            # target_cell.protection = source_cell.protection.copy()
            # target_cell.alignment = source_cell.alignment.copy()
            # Direct assignment for style attributes, using the copy module for deep copying if necessary
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy.copy(source_cell.protection)
            target_cell.alignment = copy.copy(source_cell.alignment)

        # Copy font style
        if source_cell.font:
            target_cell.font = Font(name=source_cell.font.name,
                                    size=source_cell.font.size,
                                    bold=source_cell.font.bold,
                                    italic=source_cell.font.italic,
                                    vertAlign=source_cell.font.vertAlign,
                                    underline=source_cell.font.underline,
                                    strike=source_cell.font.strike,
                                    color=source_cell.font.color)

        # Copy border style
        if source_cell.border:
            target_cell.border = Border(left=source_cell.border.left,
                                        right=source_cell.border.right,
                                        top=source_cell.border.top,
                                        bottom=source_cell.border.bottom,
                                        diagonal=source_cell.border.diagonal,
                                        diagonal_direction=source_cell.border.diagonal_direction,
                                        outline=source_cell.border.outline)

        # Copy fill style
        if source_cell.fill:
            target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                           start_color=source_cell.fill.start_color,
                                           end_color=source_cell.fill.end_color)

        # Copy alignment style
        if source_cell.alignment:
            target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                              vertical=source_cell.alignment.vertical,
                                              text_rotation=source_cell.alignment.text_rotation,
                                              wrap_text=source_cell.alignment.wrap_text,
                                              shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                              indent=source_cell.alignment.indent)

        # Copy number format, check if it's not None
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format if source_cell.number_format != 'General' else 'General'

        # Copy protection
        if source_cell.protection:
            target_cell.protection = Protection(locked=source_cell.protection.locked,
                                                hidden=source_cell.protection.hidden)

    def copy_cell(self, source_cell, target_cell, source_ws, target_ws):
        # First, check if the cell is part of a merge range and if it's not the top-left cell
        if source_ws.merged_cells.ranges:
            for range_ in source_ws.merged_cells.ranges:
                if source_cell.coordinate in range_ and source_cell.coordinate != range_.start_cell.coordinate:
                    return  # Skip copying values/styles for non-top-left cells in merged ranges

        # Copy the cell value
        target_cell.value = source_cell.value
        # Copy the style
        self.copy_style(source_cell, target_cell)

    def copy_dimensions_and_merged_cells(self, source_ws, target_ws):
        # Copy column widths
        for col in source_ws.columns:
            target_ws.column_dimensions[get_column_letter(col[0].column)].width = source_ws.column_dimensions[
                get_column_letter(col[0].column)].width

        # Copy row heights
        for row in source_ws.rows:
            target_ws.row_dimensions[row[0].row].height = source_ws.row_dimensions[row[0].row].height

        # Copy merged cell ranges
        for range_ in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(range_))

    def is_generic_name(self, sheet_name):
        return sheet_name.strip().lower().startswith("sheet")

    def get_unique_sheet_name(self, base_name, existing_names):
        if base_name not in existing_names:
            return base_name
        else:
            suffix = 1
            while f"{base_name}_{suffix}" in existing_names:
                suffix += 1
            return f"{base_name}_{suffix}"

    def combine_excel_files(self):
        file_paths = filedialog.askopenfilenames(title="Select Excel files", filetypes=[("Excel files", "*.xlsx")])
        if file_paths:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            output_path = os.path.expanduser(f'~/Downloads/combined_excel_{timestamp}.xlsx')

            target_wb = Workbook()
            target_wb.remove(target_wb.active)  # Remove the default sheet
            existing_sheet_names = set()

            for file_path in file_paths:
                source_wb = load_workbook(file_path, data_only=True)
                file_base_name = os.path.splitext(os.path.basename(file_path))[0][:10]
                for sheet_name in source_wb.sheetnames:
                    source_ws = source_wb[sheet_name]

                    if self.is_generic_name(sheet_name):
                        modified_sheet_name = self.get_unique_sheet_name(file_base_name, existing_sheet_names)
                    else:
                        modified_sheet_name = self.get_unique_sheet_name(sheet_name, existing_sheet_names)

                    existing_sheet_names.add(modified_sheet_name)
                    target_ws = target_wb.create_sheet(title=modified_sheet_name)

                    # Copy merged cells dimensions
                    self.copy_dimensions_and_merged_cells(source_ws, target_ws)

                    # Copy column widths
                    for col in source_ws.columns:
                        target_ws.column_dimensions[get_column_letter(col[0].column)].width = \
                            source_ws.column_dimensions[get_column_letter(col[0].column)].width

                    # Copy row heights
                    for row in source_ws.rows:
                        target_ws.row_dimensions[row[0].row].height = source_ws.row_dimensions[row[0].row].height

                    for row in source_ws.iter_rows():
                        for cell in row:
                            target_cell = target_ws.cell(row=cell.row, column=cell.column)
                            # self.copy_cell(cell, target_cell)
                            self.copy_cell(cell, target_cell, source_ws, target_ws)

            target_wb.save(output_path)
            messagebox.showinfo("Success", f"Excel files have been successfully combined.\n"
                                           f"{os.path.basename(output_path)} has been saved in your Downloads folder.")


def main():
    root = tk.Tk()
    # Set window size
    window_width = 800
    window_height = 400

    # Get screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calculate position for the window to be centered
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)

    root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
    app = App(root)

    root.mainloop()


if __name__ == "__main__":
    main()
